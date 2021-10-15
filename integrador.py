
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import requests
 
##########################
 
def guardar(orden):
	ws = wb.active
	ws.append(orden)
	wb.save(filename = "empanadas.xlsx" )
	f = open("datostarde.csv","a")
	f.write(orden[0]+","+str(orden[1])+","+str(orden[2])+","+str(orden[3])+","+str(orden[4])+"\n")
	f.close()
 
def validar(dato):
	try:
		dato = int(dato)
	except ValueError:
		dato = -1
	return dato
 
 
def borrar():
	ccarne.delete(0,tk.END)
	cjyq.delete(0,tk.END)
	ch.delete(0,tk.END)
	ccliente.delete(0,tk.END)
 
 
def pedido():
	cantc = ccarne.get()
	cantc = validar(cantc)
	cantjyq = cjyq.get()
	cantjyq = validar(cantjyq)
	canth = ch.get()
	canth = validar(canth)
	cunit = costos()
	if cantc>=0 and cantjyq>=0 and canth>=0:
		cliente = ccliente.get()
		if cliente:
			respuesta = messagebox.askyesno(title="Pregunta", message="¿Confirma el pedido?")
			if respuesta:
				costot = (cantc+cantjyq+canth)*cunit
				gustos = [cliente,cantc,cantjyq,canth,costot]
				#guardar
				guardar(gustos)
				messagebox.showinfo(title="Información", message="Pedido Exitoso")
				borrar()
			else:
				messagebox.showinfo(title="Información", message="Pedido en pausa")
		else:
			messagebox.showwarning(title="Advertencia", message="Error, ingrese nombre del cliente")
	else:
		messagebox.showwarning(title="Advertencia", message="Error, ingrese datos correctos")
 
def cancelar_pedido():
	respuesta = messagebox.askyesno(title="Pregunta", message="¿Desea cancelar el pedido?")
	if respuesta:
		borrar()
 
def comprobarArchivo():
	existe = os.path.exists("empanadas.xlsx")
	if existe:
		wb = load_workbook(filename = "empanadas.xlsx")
		ws = wb.active
	else:
		wb = Workbook()
		ws = wb.active
		titulo = ["Nombre","Carne","JYQ","Humita","Total"]
		ws.append(titulo)
		wb.save(filename = "empanadas.xlsx" )
		f = open("datostarde.csv","a")
		f.write("Nombre,Carne,JYQ,Humita,Total\n")
		f.close()
	return wb
 
def costos():
	r = requests.get("https://www.dolarsi.com/api/api.php?type=cotizador")
	valor = r.json()[0]["casa"]["venta"]
	dolar = float(valor.replace(",","."))
	costo = round(dolar/10) * 10
	return costo
 
 
 
##########################
 
wb = comprobarArchivo()
 
##########################
 
ventana = tk.Tk()
ventana.config(width = 400, height = 400)
ventana.title("Delivery")
 
#####etiquetas######
ebienvenido = tk.Label(text="------ Pedidos -------")
ebienvenido.place(x = 140, y = 20)
ecarne = tk.Label(text = "Ingrese cantidad de EC : ")
ecarne.place(x = 50, y = 90)
ejyq = tk.Label(text = "Ingrese cantidad de JYQ : ")
ejyq.place(x = 50, y = 150)
eh = tk.Label(text = "Ingrese cantidad de H : ")
eh.place(x = 50, y = 210)
ecliente = tk.Label(text = "Nombre del cliente : ")
ecliente.place(x = 50, y = 280)
 
 
 
#####cajas#########
ccarne = tk.Entry()
ccarne.place(x = 230, y = 90)
cjyq = tk.Entry()
cjyq.place(x = 230, y = 150)
ch = tk.Entry()
ch.place(x = 230, y = 210)
ccliente = tk.Entry()
ccliente.place(x = 230, y = 280)
 
##### Botones #########
 
bpedido = tk.Button(text = "Hacer Pedido", command = pedido)
bpedido.place(x = 270 , y = 330, height=40, width = 100)
 
bcancelar = tk.Button(text = "Cancelar Pedido", command = cancelar_pedido)
bcancelar.place(x = 150 , y = 330, height=40, width = 100)
 
binfo = tk.Button(text = "Info")
binfo.place(x = 30 , y = 330, height=40, width = 100)
 
 
ventana.mainloop()




